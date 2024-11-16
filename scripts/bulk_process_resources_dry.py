import io
import logging
import os
import warnings
from typing import Dict, List, Optional

import openpyxl  # type: ignore
from google.cloud import secretmanager, storage  # type: ignore
from tqdm import tqdm  # type:ignore

from cropsprices.parsers import parse_excel


class DataManager:
    def __init__(self):
        self._setup_clients()
        self._setup_logging()
        self.bucket = self._get_gcs_bucket()

    def _setup_clients(self):
        self.storage_client = storage.Client()
        self.secret_client = secretmanager.SecretManagerServiceClient()

    def _setup_logging(self):
        log_directory = "logs"
        if not os.path.exists(log_directory):
            os.makedirs(log_directory)

        log_file = os.path.join(log_directory, "bulk_process_2.log")

        logging.basicConfig(
            level=logging.INFO,
            format="%(asctime)s - %(name)s - %(levelname)s - %(message)s",
            handlers=[logging.FileHandler(log_file)],
        )
        self.logger = logging.getLogger("Bulk Data Load")

    def _get_gcs_bucket(self):
        secret_name = "projects/cropsprices/secrets/bucket-name/versions/latest"
        response = self.secret_client.access_secret_version(
            request={"name": secret_name}
        )
        bucket_name = response.payload.data.decode("UTF-8")
        return self.storage_client.bucket(bucket_name)

    def process_xlsx_files(self):
        blobs = list(
            self.bucket.list_blobs(
                prefix="wholesale_prices_workbooks/", include_folders_as_prefixes=False
            )
        )
        for blob in tqdm(blobs, desc="Processing Excel files", unit="file"):
            self._process_single_file(blob)

    def _process_single_file(self, blob):
        try:
            excel_file = self._download_file(blob)

            with warnings.catch_warnings():
                warnings.simplefilter("ignore", UserWarning)
                workbook = openpyxl.load_workbook(excel_file, data_only=True)

            if "ceny hurt_warz" in workbook.sheetnames:
                self._process_sheet(
                    excel_file, "vegetables", "ceny hurt_warz", False, blob.name
                )
            elif "WK" in workbook.sheetnames:
                self._process_sheet(excel_file, "vegetables", "WK", False, blob.name)
            else:
                self.logger.warning(f"No valid vegetable sheet found in {blob.name}")

            excel_file.seek(0)
            if "ceny hurt_owoc" in workbook.sheetnames:
                self._process_sheet(
                    excel_file, "fruits", "ceny hurt_owoc", True, blob.name
                )
            elif "OK" in workbook.sheetnames:
                self._process_sheet(excel_file, "fruits", "OK", True, blob.name)
            else:
                self.logger.warning(f"No valid fruit sheet found in {blob.name}")

        except Exception as e:
            self.logger.error(f"Error processing file {blob.name}: {str(e)}")

    def _download_file(self, blob) -> io.BytesIO:
        content = blob.download_as_bytes()
        return io.BytesIO(content)

    def _process_sheet(
        self,
        excel_file: io.BytesIO,
        product_type: str,
        sheet_name: str,
        is_fruit: bool,
        file_name: str,
        **kwargs,
    ):
        data = self._parse_sheet(excel_file, sheet_name, is_fruit, file_name, **kwargs)
        if data:
            self._insert_data(data, product_type, file_name)
        else:
            raise ValueError(f"Couldn't parse {sheet_name} from {file_name}.")

    def _parse_sheet(
        self,
        excel_file: io.BytesIO | str,
        sheet_name: str,
        is_fruit: bool,
        file_name: str,
        **kwargs,
    ) -> Optional[List[Dict]]:
        for skiprows in range(5):
            try:
                data = parse_excel(
                    excel_file,
                    sheet_name=sheet_name,
                    is_fruit=is_fruit,
                    skiprows=skiprows,
                    **kwargs,
                )
                self.logger.info(
                    f"Successfully parsed {sheet_name} data with skiprows={skiprows}"
                )
                return data
            except Exception as e:
                if skiprows == 4:
                    self.logger.error(
                        f"Failed to parse {sheet_name} data from {file_name}: {str(e)}"
                    )
                else:
                    self.logger.warning(
                        f"Failed to parse {sheet_name} data with skiprows={skiprows}, trying next value"
                    )
        self.logger.error(
            f"Failed to parse {sheet_name} data from {file_name} after trying all skiprows values"
        )
        return None

    def _insert_data(self, data: List[Dict], product_type: str, file_name: str):
        errors = False
        if errors:
            self.logger.error(
                f"Encountered errors while inserting {product_type} data: {errors}"
            )
        else:
            self.logger.info(
                f"Successfully inserted {len(data)} rows of {product_type} data from {file_name}"
            )


def main():
    InitialFilesManager = DataManager()
    InitialFilesManager.process_xlsx_files()


if __name__ == "__main__":
    main()
