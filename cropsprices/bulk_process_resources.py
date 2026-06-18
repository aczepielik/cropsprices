import io
import logging
import warnings
from pathlib import Path
from typing import Optional

import openpyxl
import pandas as pd
from tqdm import tqdm

from cropsprices.parsers import parse_excel

logging.basicConfig(level=logging.INFO, format="%(asctime)s - %(levelname)s - %(message)s")
logger = logging.getLogger("Bulk Data Load")


class DataManager:
    def __init__(self, raw_dir: str = "data/raw", parsed_dir: str = "data/parsed"):
        self.raw_dir = Path(raw_dir)
        self.parsed_dir = Path(parsed_dir)
        self.raw_dir.mkdir(parents=True, exist_ok=True)
        self.parsed_dir.mkdir(parents=True, exist_ok=True)

    def process_xlsx_files(self):
        xlsx_files = sorted(self.raw_dir.glob("*.xlsx"))
        logger.info(f"Found {len(xlsx_files)} XLSX files in {self.raw_dir}")

        for filepath in tqdm(xlsx_files, desc="Processing Excel files", unit="file"):
            self._process_single_file(filepath)

    VEG_SHEET_NAMES = ["ceny hurt_warz", "HURT WARZ", "WK"]
    FRUIT_SHEET_NAMES = ["ceny hurt_owoc", "HURT OWOC", "OK"]

    def _process_single_file(self, filepath: Path):
        try:
            with open(filepath, "rb") as f:
                excel_bytes = io.BytesIO(f.read())

            with warnings.catch_warnings():
                warnings.simplefilter("ignore", UserWarning)
                workbook = openpyxl.load_workbook(excel_bytes, data_only=True)

            file_stem = filepath.stem

            veg_sheet = next(
                (s for s in self.VEG_SHEET_NAMES if s in workbook.sheetnames), None
            )
            if veg_sheet:
                self._process_sheet(excel_bytes, "vegetables", veg_sheet, False, file_stem)
            else:
                logger.warning(f"No valid vegetable sheet found in {filepath.name}")

            excel_bytes.seek(0)
            fruit_sheet = next(
                (s for s in self.FRUIT_SHEET_NAMES if s in workbook.sheetnames), None
            )
            if fruit_sheet:
                self._process_sheet(excel_bytes, "fruits", fruit_sheet, True, file_stem)
            else:
                logger.warning(f"No valid fruit sheet found in {filepath.name}")

        except Exception as e:
            logger.error(f"Error processing file {filepath.name}: {str(e)}")

    def _process_sheet(
        self,
        excel_file: io.BytesIO,
        product_type: str,
        sheet_name: str,
        is_fruit: bool,
        file_stem: str,
    ):
        data = self._parse_sheet(excel_file, sheet_name, is_fruit, file_stem)
        if data:
            output_path = self.parsed_dir / f"{file_stem}_{product_type}.csv"
            df = pd.DataFrame(data)
            df.to_csv(output_path, index=False)
            logger.info(f"Saved {len(data)} rows to {output_path.name}")
        else:
            raise ValueError(f"Couldn't parse {sheet_name} from {file_stem}.")

    def _parse_sheet(
        self,
        excel_file: io.BytesIO,
        sheet_name: str,
        is_fruit: bool,
        file_stem: str,
    ) -> Optional[list]:
        for skiprows in range(5):
            try:
                data = parse_excel(
                    excel_file,
                    sheet_name=sheet_name,
                    is_fruit=is_fruit,
                    skiprows=skiprows,
                )
                logger.info(
                    f"Successfully parsed {sheet_name} data with skiprows={skiprows}"
                )
                return data
            except Exception as e:
                if skiprows == 4:
                    logger.error(
                        f"Failed to parse {sheet_name} data from {file_stem}: {str(e)}"
                    )
                else:
                    logger.warning(
                        f"Failed to parse {sheet_name} data with skiprows={skiprows}, trying next value"
                    )
        logger.error(
            f"Failed to parse {sheet_name} data from {file_stem} after trying all skiprows values"
        )
        return None


def main():
    data_manager = DataManager()
    data_manager.process_xlsx_files()


if __name__ == "__main__":
    main()
